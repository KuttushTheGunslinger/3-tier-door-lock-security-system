import smtplib
import random
from datetime import *
import cv2
import numpy
import os
from openpyxl import load_workbook
from pytz import timezone

wb = load_workbook('personal_data.xlsx')
ws = wb.active

k = 0
j = 0
k1 = 0
j1 = 0

haar = 'haarcascade_frontalface_default.xml'
data = ['level_1', 'level_2']

print('Recognizing face please be in sufficient lights...')

(images, lables, names, id) = ([], [], {}, 0)

for d in data:
    for (subdirs, dirs, files) in os.walk(d):
        for subdir in dirs:
            names[id] = subdir
            subjectpath = os.path.join(d, subdir)
            for filename in os.listdir(subjectpath):
                path = subjectpath + '/' + filename
                lable = id
                images.append(cv2.imread(path, 0))
                lables.append(int(lable))
            id += 1
(width, height) = (130, 100)
(images, lables) = [numpy.array(lis) for lis in [images, lables]]

model = cv2.face.LBPHFaceRecognizer_create()
model.train(images, lables)

face_cascade = cv2.CascadeClassifier(haar)
webcam = cv2.VideoCapture(0)

while True:
    (_, im) = webcam.read()
    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    for (x, y, w, h) in faces:
        cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 2)
        face = gray[y:y + h, x:x + w]
        face_resize = cv2.resize(face, (width, height))
        pred = model.predict(face_resize)
        cv2.rectangle(im, (x, y), (x + w, y + h), (0, 255, 0), 3)

        if pred[1] < 100:
            now = datetime.now(timezone('UTC'))
            now_asia = now.astimezone(timezone('Asia/Kolkata'))

            if names[pred[0]] in os.listdir('level_1'):
                a = 1
                for cell in ws['A']:
                    if cell.value is not None:
                        if names[pred[0]] in cell.value:
                            r = cell.row

                r1 = 'B' + str(r)
                r2 = 'C' + str(r)

                while k < 3:
                    p = input('Enter your password: ')

                    if p == ws[r1].value:
                        otp = random.randint(1000, 9999)

                        server = smtplib.SMTP('smtp.gmail.com', port=587)
                        server.starttls()
                        server.login("gajeelredfox746@gmail.com", "passWORD123")

                        msg = str(otp)

                        server.sendmail("gajeelredfox746@gmail.com", ws[r2].value, msg)

                        server.quit()

                        print('Check your mail for OTP')

                        otp1 = int(input('Enter the OTP: '))

                        while j < 3:
                            if otp1 == otp:
                                print('Level 1 access granted')
                                print('Welcome ', names[pred[0]])

                                wb1 = load_workbook('entry_log.xlsx')
                                ws1 = wb1.active

                                r2 = ws1.max_row + 1
                                ws1.cell(column=1, row=r2, value=names[pred[0]])
                                ws1.cell(column=2, row=r2, value=now_asia.strftime("%d/%m/%Y"))
                                ws1.cell(column=3, row=r2, value=now_asia.strftime("%H:%M:%S"))

                                wb1.save(filename='entry_log.xlsx')
                                wb1.close()
                                break

                            else:
                                print('Incorrect OTP')
                                j += 1

                        break

                    else:
                        print('Incorrect Password')
                        k += 1

            else:
                a = 2
                for cell in ws['A']:
                    if cell.value is not None:
                        if names[pred[0]] in cell.value:
                            r = cell.row

                r1 = 'B' + str(r)
                r2 = 'C' + str(r)

                while k1 < 3:
                    p = input('Enter your password: ')

                    if p == ws[r1].value:
                        print('Level 2 access granted')
                        print('Welcome ', names[pred[0]])

                        wb1 = load_workbook('entry_log.xlsx')
                        ws1 = wb1.active

                        r2 = ws1.max_row + 1
                        ws1.cell(column=1, row=r2, value=names[pred[0]])
                        ws1.cell(column=2, row=r2, value=now_asia.strftime("%d/%m/%Y"))
                        ws1.cell(column=3, row=r2, value=now_asia.strftime("%H:%M:%S"))

                        wb1.save(filename='entry_log.xlsx')
                        wb1.close()
                        break

                    else:
                        print('Incorrect Password')
                        k1 += 1

            cv2.putText(im, '%s - level %s- %.0f' % (names[pred[0]], a, pred[1]), (x - 10, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))

        else:
            cv2.putText(im, 'not recognized', (x - 10, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 0))
            print('Not recognized. Please try again.')

    # cv2.imshow('OpenCV', im)
    #
    # key = cv2.waitKey(10)
    # if key == 27:
    #     break

